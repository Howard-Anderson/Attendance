"""
            Attendance Maintainer: Face Detection Using CNN:

    Author: Howard Anderson.

    Date: 06/02/2024.

    Description: Module to classify faces and enter the Attendance.

    Filename: attendance.py.
"""

# Importing Modules.
import os
import cv2
import numpy as np
import pandas as pd
from datetime import datetime
from openpyxl import workbook, load_workbook

# Imports from Tensorflow.
from tensorflow.keras.models import load_model

# Capture Face Image.
def capture_image():
    cap = cv2.VideoCapture(0)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 480)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

    ret, frame = cap.read()

    cv2.imshow("Image", frame)
    cv2.waitKey(0)

    cap.release()
    cv2.destroyAllWindows()

    return frame

# Preprocess Image.
def preprocess_image(image):
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    img_array = np.expand_dims(img, axis = 0)
    return img_array

# Update Attendance in Excel Sheets.
def update_attendance(name):
    filename = "Attendance.xlsx"
    date_time = datetime.now().strtime("%Y-%m-%d %H:%M:%S")

    if(os.path.exists(filename)):
        wb = load_workbook(filename)
    else:
        wb = workbook()

    if name not in wb.sheetnames:
        wb.create_sheetname(name)
        ws = wb[name]
        ws.append(["Date", "Time"])
    else:
        ws = wb[name]

    ws.append([
        date_time.split()[0],
        date_time.split()[1])
    ])

    wb.save(filename)


if __name__ == "__main__":
    input_image = capture_image()
    input_image = preprocess_image(input_image)

    classifier = load_model("Classifier.h5")
    pred = classifier.predict(input_image)
    pred = np.argmax(pred)
    df_classnames = pd.read_csv("lookup.csv")
    name = df_classnames[0][pred]
    update_attendance(pred)




